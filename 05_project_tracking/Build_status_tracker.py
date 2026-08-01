"""
Builds status_tracker.xlsx showing the project lifecycle phases,
owners, and status, per BRD milestones.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Status Tracker"

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
DONE_FONT = Font(name="Arial", color="006100")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

headers = ["Phase", "Task", "Owner", "Target Date", "Status", "Notes"]
col_widths = [22, 32, 22, 15, 12, 40]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP
    cell.border = BORDER

tasks = [
    ["Requirements", "Draft Business Requirements Document", "Business/Operations Analyst", "Week 1", "Done",
     "BRD covers scope, stakeholders, functional requirements, and acceptance criteria"],
    ["Requirements", "Circulate BRD for stakeholder review", "Business/Operations Analyst", "Week 1", "Done",
     "Reviewed by Operations, Client Relationship, and Compliance"],
    ["Requirements", "Stakeholder sign-off on requirements", "All Stakeholders", "Week 1", "Done",
     "Approved by all three stakeholder groups"],
    ["Data Setup", "Build mock source dataset", "Business/Operations Analyst", "Week 2", "Done",
     "25 mock accounts, 23 active / 2 closed, generated with Faker"],
    ["Build", "Develop report generation pipeline", "Business/Operations Analyst", "Week 2", "Done",
     "Python/Pandas pipeline built against FR-01 through FR-06"],
    ["Build", "Internal review of generated output", "Business/Operations Analyst", "Week 2", "Done",
     "Spot-checked balances and account counts before UAT"],
    ["UAT", "Write UAT test cases", "Business/Operations Analyst", "Week 3", "Done",
     "10 test cases mapped to functional requirements and data quality scenarios"],
    ["UAT", "Execute UAT test cases", "Business/Operations Analyst", "Week 3", "Done",
     "All 10 test cases passed; results logged in UAT_Test_Cases.xlsx"],
    ["UAT", "Log and resolve any defects", "Business/Operations Analyst", "Week 3", "Done",
     "No defects found in this run"],
    ["Sign-off", "Final stakeholder review of tested report", "All Stakeholders", "Week 4", "Done",
     "Report approved for use as the reference deliverable"],
    ["Sign-off", "Archive BRD, UAT results, and sample output", "Business/Operations Analyst", "Week 4", "Done",
     "All artifacts stored in project repository"],
]

for row_idx, row_data in enumerate(tasks, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if col_idx == 5:  # Status column
            cell.fill = DONE_FILL
            cell.font = DONE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 22

wb.save("status_tracker.xlsx")
print("Saved status_tracker.xlsx")

